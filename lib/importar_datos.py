# -*- coding: utf-8 -*-
"""Importación automática de indicadores desde Excel, CSV o Google Sheet."""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from lib.mdeia_model import (
    indicadores_df,
    load_framework,
    load_indicadores,
    normalizar_respuestas,
    pilot_codigos,
)
from lib.oia_encuesta import fetch_sheet_csv
from lib.texto_legible import legibilizar_siglas_udigital

_CODIGO_HINTS = ("codigo", "código", "code", "id_indicador")
_VALOR_HINTS = ("valor", "value", "respuesta", "nivel", "puntaje", "resultado")
_SI_NO_OPCIONES = [
    "Sí — cumple / existe",
    "No — no cumple / no existe",
]


def _nivel_opciones() -> list[str]:
    fw = load_framework()
    return [f"{n['valor']} — {n['etiqueta']}" for n in fw["niveles_madurez"]]


def _norm_col(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip().lower())


def _find_column(columns: list[str], hints: tuple[str, ...]) -> str | None:
    normalized = {_norm_col(c): c for c in columns}
    for h in hints:
        if h in normalized:
            return normalized[h]
    for nc, orig in normalized.items():
        if any(h in nc for h in hints):
            return orig
    return None


def _strip_etiqueta_valor(text: str) -> str:
    """Quita la etiqueta descriptiva de opciones del desplegable (p. ej. «3 — Implementado»)."""
    for sep in (" — ", " – ", " - "):
        if sep in text:
            return text.split(sep, 1)[0].strip()
    return text.strip()


def _texto_ayuda_valor(ind: dict) -> str:
    tipo = ind.get("tipo") or "nivel"
    if tipo == "nivel":
        return (
            "Elegir en el desplegable el nivel 0–4 según madurez "
            "(consensuado con evidencia; no ingresar % ni Sí/No crudos)."
        )
    if tipo == "si_no":
        return "Elegir Sí o No según exista la política, guía o práctica indicada."
    if tipo == "porcentaje":
        umbral = ind.get("umbral", 60)
        return f"Ingresar número 0–100 (%). Referencia de satisfacción del IMD: ≥ {umbral} %."
    if tipo == "umbral":
        umbral = ind.get("umbral", 70)
        return f"Ingresar número 0–100 (%). Referencia de satisfacción del IMD: ≥ {umbral} %."
    return "Completar según el indicador."


def _parse_valor(raw: Any, tipo: str) -> Any:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    text = str(raw).strip()
    if not text:
        return None
    text = _strip_etiqueta_valor(text)
    if tipo == "si_no":
        low = text.lower()
        if low in {"sí", "si", "yes", "true", "1", "s"}:
            return "Sí"
        if low in {"no", "false", "0", "n"}:
            return "No"
        return text
    if tipo in {"porcentaje", "umbral"}:
        try:
            return float(text.replace(",", ".").replace("%", ""))
        except ValueError:
            return None
    try:
        if "." in text or "," in text:
            return float(text.replace(",", "."))
    except ValueError:
        pass
    try:
        return int(float(text))
    except ValueError:
        return text


def parse_dataframe_indicadores(df: pd.DataFrame) -> dict[str, Any]:
    """Convierte tabla codigo/valor en dict de respuestas MDeIA."""
    if df.empty:
        return {}
    col_cod = _find_column(list(df.columns), _CODIGO_HINTS)
    col_val = _find_column(list(df.columns), _VALOR_HINTS)
    if not col_cod or not col_val:
        raise ValueError(
            "La planilla debe tener columnas **codigo** y **valor** "
            "(o equivalentes: código, respuesta, nivel)."
        )

    meta = {ind["codigo"]: ind for ind in load_indicadores()}
    out: dict[str, Any] = {}
    for _, row in df.iterrows():
        cod = str(row[col_cod]).strip().upper()
        if not cod or cod == "NAN":
            continue
        cod = cod.replace("CLJL_IA_", "MDEIA_IA_")
        ind = meta.get(cod)
        if not ind:
            continue
        val = _parse_valor(row[col_val], ind.get("tipo", "nivel"))
        if val is not None:
            out[cod] = val
    return normalizar_respuestas(out)


def _leer_planilla_excel(data: bytes) -> pd.DataFrame:
    """Lee la hoja con codigo/valor (Indicadores) aunque no sea la primera pestaña."""
    xl = pd.ExcelFile(BytesIO(data))
    if "Indicadores" in xl.sheet_names:
        return pd.read_excel(xl, sheet_name="Indicadores")
    for name in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=name)
        if _find_column(list(df.columns), _CODIGO_HINTS) and _find_column(
            list(df.columns), _VALOR_HINTS
        ):
            return df
    return pd.read_excel(xl, sheet_name=xl.sheet_names[0])


def load_uploaded_indicadores(name: str, data: bytes) -> dict[str, Any]:
    lower = name.lower()
    if lower.endswith((".xlsx", ".xls")):
        df = _leer_planilla_excel(data)
    elif lower.endswith(".csv"):
        df = pd.read_csv(BytesIO(data))
    else:
        raise ValueError("Formato no soportado. Usá .xlsx o .csv")
    return parse_dataframe_indicadores(df)


def parse_google_sheet_url(text: str) -> tuple[str, str]:
    """Extrae ID y GID desde URL completa o ID suelto."""
    raw = text.strip()
    if not raw:
        raise ValueError("Ingresá la URL o el ID de la planilla.")
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", raw)
    if m:
        sheet_id = m.group(1)
    elif re.fullmatch(r"[a-zA-Z0-9-_]{20,}", raw):
        sheet_id = raw
    else:
        raise ValueError(
            "No reconocí el ID. Pegá la URL completa de Google Sheets o solo el ID."
        )
    gid_m = re.search(r"[?&#]gid=(\d+)", raw)
    gid = gid_m.group(1) if gid_m else "0"
    return sheet_id, gid


def load_sheet_indicadores(sheet_id: str, gid: str | int = "0") -> dict[str, Any]:
    df = fetch_sheet_csv(sheet_id, gid)
    return parse_dataframe_indicadores(df)


def _row_ranges(rows: list[int]) -> list[tuple[int, int]]:
    if not rows:
        return []
    ordered = sorted(rows)
    ranges: list[tuple[int, int]] = []
    start = end = ordered[0]
    for row in ordered[1:]:
        if row == end + 1:
            end = row
        else:
            ranges.append((start, end))
            start = end = row
    ranges.append((start, end))
    return ranges


def _escribir_hoja_escalas(writer: pd.ExcelWriter) -> None:
    fw = load_framework()
    niveles = fw["niveles_madurez"]
    filas = []
    for n in niveles:
        sat = "satisface IMD" if n.get("satisface") else "no satisface IMD"
        filas.append(
            {
                "Valor": n["valor"],
                "Significado": n["etiqueta"],
                "Para el IMD": sat,
            }
        )
    df_n = pd.DataFrame(filas)
    df_sn = pd.DataFrame(
        {
            "Opción desplegable": _SI_NO_OPCIONES,
            "Significado": [
                "La política, guía o práctica existe y está vigente.",
                "No existe o no está formalizada.",
            ],
        }
    )
    df_n.to_excel(writer, index=False, sheet_name="Escalas", startrow=0)
    ws = writer.sheets["Escalas"]
    ws.cell(row=df_n.shape[0] + 3, column=1, value="Sí / No (indicadores binarios)")
    for i, row in df_sn.iterrows():
        ws.cell(row=df_n.shape[0] + 4 + i, column=1, value=row["Opción desplegable"])
        ws.cell(row=df_n.shape[0] + 4 + i, column=2, value=row["Significado"])


def _escribir_hoja_instrucciones(writer: pd.ExcelWriter) -> None:
    instrucciones = pd.DataFrame(
        {
            "Columna": [
                "codigo",
                "indicador",
                "tipo",
                "como_completar",
                "valor",
            ],
            "Qué hacer": [
                "No modificar. Identificador del indicador MDeIA.",
                "No modificar. Enunciado de referencia.",
                "No modificar. Tipo de respuesta esperada (referencia).",
                "Leer antes de completar. Explica qué va en valor.",
                "ÚNICA columna editable. Usar desplegable o número según tipo.",
            ],
        }
    )
    reglas = pd.DataFrame(
        {
            "tipo": ["nivel", "si_no", "porcentaje", "umbral"],
            "Qué poner en valor": [
                "Desplegable 0–4 (ver hoja Escalas). Madurez consensuada, no dato crudo.",
                "Desplegable Sí / No.",
                "Número 0–100 (porcentaje real medido).",
                "Número 0–100 (porcentaje o tasa medida).",
            ],
            "Ejemplo": ["3 — Implementado", "Sí — cumple / existe", "45", "72"],
        }
    )
    instrucciones.to_excel(writer, index=False, sheet_name="Instrucciones", startrow=0)
    ws = writer.sheets["Instrucciones"]
    ws.cell(row=instrucciones.shape[0] + 3, column=1, value="Reglas por tipo")
    for i, row in reglas.iterrows():
        ws.cell(row=instrucciones.shape[0] + 4 + i, column=1, value=row["tipo"])
        ws.cell(row=instrucciones.shape[0] + 4 + i, column=2, value=row["Qué poner en valor"])
        ws.cell(row=instrucciones.shape[0] + 4 + i, column=3, value=row["Ejemplo"])


def _formatear_hoja_indicadores(writer: pd.ExcelWriter, export: pd.DataFrame) -> None:
    ws = writer.sheets["Indicadores"]
    ws_escalas = writer.sheets["Escalas"]
    nivel_opciones = _nivel_opciones()
    n_nivel = len(nivel_opciones)
    for i, opcion in enumerate(nivel_opciones, start=1):
        ws_escalas.cell(row=i, column=4, value=opcion)
    for i, opcion in enumerate(_SI_NO_OPCIONES, start=1):
        ws_escalas.cell(row=i, column=6, value=opcion)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    readonly_fill = PatternFill("solid", fgColor="F2F2F2")
    valor_fill = PatternFill("solid", fgColor="FFF9E6")
    unlock = Protection(locked=False)
    lock = Protection(locked=True)

    col_valor = list(export.columns).index("valor") + 1
    col_letter_valor = get_column_letter(col_valor)

    for col_idx, name in enumerate(export.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {"codigo": 28, "indicador": 52, "tipo": 12, "como_completar": 44, "valor": 34}
    for col_idx, name in enumerate(export.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 18)

    rows_por_tipo: dict[str, list[int]] = {"nivel": [], "si_no": [], "numero": []}
    for excel_row, tipo in enumerate(export["tipo"], start=2):
        for col_idx in range(1, len(export.columns) + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == col_valor:
                cell.fill = valor_fill
                cell.protection = unlock
            else:
                cell.fill = readonly_fill
                cell.protection = lock
        if tipo == "nivel":
            rows_por_tipo["nivel"].append(excel_row)
        elif tipo == "si_no":
            rows_por_tipo["si_no"].append(excel_row)
        elif tipo in {"porcentaje", "umbral"}:
            rows_por_tipo["numero"].append(excel_row)

    for start, end in _row_ranges(rows_por_tipo["nivel"]):
        dv = DataValidation(
            type="list",
            formula1=f"=Escalas!$D$1:$D${n_nivel}",
            allow_blank=True,
        )
        dv.error = "Elegí un nivel del desplegable (0 a 4)."
        dv.errorTitle = "Valor inválido"
        dv.prompt = "Nivel de madurez digital según evidencia institucional."
        dv.promptTitle = "Nivel 0–4"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter_valor}{start}:{col_letter_valor}{end}")

    for start, end in _row_ranges(rows_por_tipo["si_no"]):
        dv = DataValidation(
            type="list",
            formula1=f"=Escalas!$F$1:$F${len(_SI_NO_OPCIONES)}",
            allow_blank=True,
        )
        dv.error = "Elegí Sí o No del desplegable."
        dv.errorTitle = "Valor inválido"
        dv.prompt = "¿Cumple la política o práctica indicada?"
        dv.promptTitle = "Sí / No"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter_valor}{start}:{col_letter_valor}{end}")

    for start, end in _row_ranges(rows_por_tipo["numero"]):
        dv = DataValidation(
            type="decimal",
            operator="between",
            formula1="0",
            formula2="100",
            allow_blank=True,
        )
        dv.error = "Ingresá un número entre 0 y 100 (sin signo %)."
        dv.errorTitle = "Porcentaje inválido"
        dv.prompt = "Porcentaje o tasa medida (0–100)."
        dv.promptTitle = "Valor numérico"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter_valor}{start}:{col_letter_valor}{end}")

    ws.freeze_panes = "A2"
    ws.protection.sheet = True
    ws.protection.enable()


def generar_plantilla_excel(*, piloto: bool = True) -> bytes:
    """Plantilla descargable con desplegables en valor y columnas de ayuda."""
    df = indicadores_df()
    if piloto:
        codes = pilot_codigos()
        df = df[df["codigo"].isin(codes)].copy()
    df = df.sort_values("codigo")

    records = []
    for _, row in df.iterrows():
        ind = row.to_dict()
        records.append(
            {
                "codigo": ind["codigo"],
                "indicador": legibilizar_siglas_udigital(ind["texto"]),
                "tipo": ind.get("tipo") or "nivel",
                "como_completar": _texto_ayuda_valor(ind),
                "valor": "",
            }
        )
    export = pd.DataFrame(records)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export.to_excel(writer, index=False, sheet_name="Indicadores")
        _escribir_hoja_escalas(writer)
        _escribir_hoja_instrucciones(writer)
        _formatear_hoja_indicadores(writer, export)
    buf.seek(0)
    return buf.getvalue()
