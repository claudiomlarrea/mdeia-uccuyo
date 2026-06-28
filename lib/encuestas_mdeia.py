# -*- coding: utf-8 -*-
"""Encuestas MDeIA por audiencia: plantillas Excel, análisis y mapeo a indicadores."""

from __future__ import annotations

import json
import re
import time
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lib.oia_encuesta import nivel_encuesta, nivel_observatorio

_DATA = Path(__file__).resolve().parent.parent / "data"
_TIMESTAMP_HINTS = ("marca temporal", "timestamp", "fecha y hora", "fecha/hora")


def lista_unidades_mdeia() -> list[str]:
    """Etiquetas oficiales de unidades académicas (mismo catálogo que MDeIA)."""
    with (_DATA / "unidades_academicas.json").open(encoding="utf-8") as f:
        data = json.load(f)
    labels: list[str] = []
    for grupo in data.get("grupos", []):
        for u in grupo.get("unidades", []):
            label = str(u.get("label") or u.get("nombre") or u.get("id", "")).strip()
            if label and label not in labels:
                labels.append(label)
    return sorted(labels, key=lambda s: s.upper())


def load_encuestas_config() -> dict:
    with (_DATA / "encuestas_mdeia.json").open(encoding="utf-8") as f:
        return json.load(f)


def listar_audiencias() -> list[dict]:
    cfg = load_encuestas_config()
    return [
        {
            "id": k,
            "nombre": v["nombre"],
            "descripcion": v.get("descripcion", ""),
            "google_form_url": v.get("google_form_url", ""),
        }
        for k, v in cfg["audiencias"].items()
    ]


def url_formulario_google(audiencia_id: str, *, overrides: dict[str, str] | None = None) -> str:
    """URL pública del Google Form (config JSON o overrides desde secrets)."""
    if overrides and audiencia_id in overrides:
        url = str(overrides[audiencia_id]).strip()
        if url:
            return url
    aud = audiencia_cfg(audiencia_id)
    return str(aud.get("google_form_url") or "").strip()


def audiencia_cfg(audiencia_id: str) -> dict:
    cfg = load_encuestas_config()
    if audiencia_id not in cfg["audiencias"]:
        raise ValueError(f"Audiencia desconocida: {audiencia_id}")
    return cfg["audiencias"][audiencia_id]


def _normalize_header(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip().lower())


def _resolve_column(df: pd.DataFrame, columna: str) -> str | None:
    target = columna.lower()
    for c in df.columns:
        raw = str(c).strip()
        if raw == columna or raw.lower() == target:
            return c
        if target in _normalize_header(raw):
            return c
    return None


def _escala_cfg(cfg: dict, escala_id: str) -> dict | None:
    if escala_id in ("texto", "meta_sede", "meta_unidad"):
        return None
    return cfg["escalas"].get(escala_id)


def _valor_numerico(raw: Any, escala: dict) -> float | None:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    text = str(raw).strip()
    if not text:
        return None
    opciones = escala.get("opciones") or []
    valores = escala.get("valores") or list(range(len(opciones)))
    for i, op in enumerate(opciones):
        if text.lower() == str(op).lower():
            return float(valores[i])
    try:
        num = float(text.replace(",", "."))
        if num.is_integer() and 0 <= num <= 4:
            return num
    except ValueError:
        pass
    return None


def _filas_validas(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    data_cols = [c for c in df.columns if not _is_timestamp_column(c)]
    if not data_cols:
        return df.iloc[0:0]
    mask = df[data_cols].notna().any(axis=1)
    for c in data_cols:
        mask |= df[c].astype(str).str.strip().astype(bool)
    return df.loc[mask].copy()


def _is_timestamp_column(name: str) -> bool:
    lc = str(name).lower()
    return any(h in lc for h in _TIMESTAMP_HINTS)


def _escala_por_columna(cfg: dict, columna: str) -> str | None:
    for aud in cfg["audiencias"].values():
        for p in aud.get("preguntas", []):
            if p["columna"] == columna:
                return p.get("escala")
    return None


def _promedio_columnas(df: pd.DataFrame, cfg: dict, columnas: list[str]) -> float | None:
    vals: list[float] = []
    for col_id in columnas:
        escala_id = _escala_por_columna(cfg, col_id)
        if not escala_id:
            continue
        escala = _escala_cfg(cfg, escala_id)
        if not escala:
            continue
        resolved = _resolve_column(df, col_id)
        if not resolved:
            continue
        for raw in df[resolved]:
            v = _valor_numerico(raw, escala)
            if v is not None:
                vals.append(v)
    if not vals:
        return None
    return sum(vals) / len(vals)


def _redondear_nivel(prom: float) -> int:
    return max(0, min(4, int(round(prom))))


def calcular_indicadores_desde_encuesta(
    df: pd.DataFrame,
    audiencia_id: str,
    *,
    poblacion: int | None = None,
) -> dict[str, Any]:
    """Analiza respuestas y devuelve métricas + indicadores sugeridos (0–4)."""
    cfg = load_encuestas_config()
    aud = audiencia_cfg(audiencia_id)
    df = _filas_validas(df)
    n = len(df)

    tasa: float | None = None
    if poblacion and poblacion > 0:
        tasa = round(min(100.0, (n / poblacion) * 100), 1)

    indicadores: dict[str, int] = {}
    detalle: list[dict] = []

    for rule in aud.get("indicadores", []):
        codigo = rule["codigo"]
        regla = rule["regla"]
        valor: int | None = None
        nota = ""

        if regla == "tasa_respuesta_nivel":
            valor = nivel_encuesta(tasa, n_respuestas=n)
            nota = f"Tasa {tasa} % ({n} respuestas)" if tasa is not None else f"{n} respuestas"
        elif regla == "volumen_respuestas_nivel":
            valor = nivel_observatorio(n)
            nota = f"{n} respuestas válidas"
        elif regla == "directo":
            pregunta = rule["pregunta"]
            prom = _promedio_columnas(df, cfg, [pregunta])
            if prom is not None:
                valor = _redondear_nivel(prom)
                nota = f"Promedio encuesta = {prom:.2f}"
        elif regla == "promedio_preguntas":
            preguntas = rule.get("preguntas") or []
            prom = _promedio_columnas(df, cfg, preguntas)
            if prom is not None:
                valor = _redondear_nivel(prom)
                nota = f"Promedio {len(preguntas)} ítems = {prom:.2f}"

        if valor is not None:
            indicadores[codigo] = valor
            detalle.append({"codigo": codigo, "valor": valor, "nota": nota, "regla": regla})

    columnas_ok = sum(
        1
        for p in aud.get("preguntas", [])
        if p.get("escala") not in ("texto",) and _resolve_column(df, p["columna"])
    )

    return {
        "audiencia": audiencia_id,
        "audiencia_nombre": aud["nombre"],
        "n_respuestas": n,
        "poblacion_objetivo": poblacion,
        "tasa_respuesta_pct": tasa,
        "columnas_reconocidas": columnas_ok,
        "columnas_esperadas": len([p for p in aud.get("preguntas", []) if p.get("escala") != "texto"]),
        "indicadores": indicadores,
        "detalle_indicadores": detalle,
        "fecha_analisis": time.strftime("%Y-%m-%dT%H:%M:%S"),
    }


def aplicar_encuesta_a_diagnostico(
    resp: dict[str, Any],
    metricas: dict[str, Any],
    *,
    sobrescribir: bool = True,
) -> dict[str, Any]:
    """Fusiona indicadores derivados de encuesta en el dict de respuestas activas."""
    out = dict(resp)
    nuevos = metricas.get("indicadores") or {}
    for codigo, valor in nuevos.items():
        if sobrescribir or codigo not in out:
            out[codigo] = valor
    return out


def resumen_encuesta_markdown(metricas: dict[str, Any]) -> str:
    lines = [
        f"- **Audiencia:** {metricas.get('audiencia_nombre', '—')}",
        f"- **Respuestas válidas:** {metricas.get('n_respuestas', 0)}",
    ]
    if metricas.get("poblacion_objetivo"):
        lines.append(f"- **Población convocada:** {metricas['poblacion_objetivo']}")
    if metricas.get("tasa_respuesta_pct") is not None:
        lines.append(f"- **Tasa de respuesta:** {metricas['tasa_respuesta_pct']} %")
    lines.append(
        f"- **Columnas reconocidas:** {metricas.get('columnas_reconocidas', 0)}"
        f" / {metricas.get('columnas_esperadas', 0)}"
    )
    ind = metricas.get("indicadores") or {}
    if ind:
        lines.append("- **Indicadores sugeridos:**")
        for cod, val in sorted(ind.items()):
            lines.append(f"  - `{cod}` → nivel **{val}**")
    return "\n".join(lines)


def generar_plantilla_xlsx(audiencia_id: str) -> bytes:
    """Plantilla Excel vacía compatible con export Google Forms."""
    cfg = load_encuestas_config()
    aud = audiencia_cfg(audiencia_id)
    preguntas = aud.get("preguntas", [])

    headers = ["Marca temporal"] + [p["columna"] for p in preguntas]
    df_resp = pd.DataFrame(columns=headers)

    filas_escalas = []
    for escala_id, escala in cfg.get("escalas", {}).items():
        for op in escala.get("opciones", []):
            filas_escalas.append({"escala": escala_id, "opcion": op})
    df_escalas = pd.DataFrame(filas_escalas)

    filas_inst = [
        {"campo": "Audiencia", "valor": aud["nombre"]},
        {"campo": "Versión", "valor": cfg.get("version", "1.0.0")},
        {"campo": "Instrucciones", "valor": cfg.get("instrucciones_generales", "")},
        {"campo": "Exportar Forms", "valor": "Google Forms → Respuestas → Descargar .xlsx"},
        {"campo": "Cargar MDeIA", "valor": "Encuestas por audiencia → Cargar respuestas"},
        {"campo": "Unidades MDeIA", "valor": "; ".join(lista_unidades_mdeia())},
    ]
    for p in preguntas:
        filas_inst.append(
            {
                "campo": p["columna"],
                "valor": f"{p['texto']} [{p.get('escala', 'texto')}]",
            }
        )
    df_inst = pd.DataFrame(filas_inst)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_inst.to_excel(writer, index=False, sheet_name="Instrucciones")
        df_escalas.to_excel(writer, index=False, sheet_name="Escalas")
        df_resp.to_excel(writer, index=False, sheet_name="Respuestas")
        _formatear_hoja_plantilla(writer, headers)
    buf.seek(0)
    return buf.getvalue()


def _formatear_hoja_plantilla(writer: pd.ExcelWriter, headers: list[str]) -> None:
    ws = writer.sheets["Respuestas"]
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(col_idx)].width = 22 if col_idx == 1 else 28

    ws_i = writer.sheets["Instrucciones"]
    ws_i.column_dimensions["A"].width = 28
    ws_i.column_dimensions["B"].width = 72


def generar_apps_script_google_forms() -> str:
    """Genera script de Google Apps Script para crear los 4 formularios."""
    cfg = load_encuestas_config()
    unidades_js = json.dumps(lista_unidades_mdeia(), ensure_ascii=False)
    lines = [
        "/**",
        " * MDeIA UCCuyo — Generador de encuestas Google Forms",
        " * Uso: script.google.com → Nuevo proyecto → pegar → ejecutar crearTodasLasEncuestasMDeIA",
        " * Requiere autorizar Google Forms. Al final verás en Registro los enlaces para compartir.",
        " */",
        "",
        f"var UNIDADES_MDEIA = {unidades_js};",
        "",
        "function crearTodasLasEncuestasMDeIA() {",
        "  var enlaces = [];",
    ]
    for aud_id in cfg["audiencias"]:
        lines.append(f"  enlaces.push(crearFormulario_{aud_id}());")
    lines.extend(
        [
            "  Logger.log(enlaces.join('\\n'));",
            "  return enlaces;",
            "}",
            "",
        ]
    )

    for aud_id, aud in cfg["audiencias"].items():
        titulo = f"MDeIA UCCuyo — Encuesta {aud['nombre']}"
        desc = aud.get("descripcion", "")
        lines.append(f"function crearFormulario_{aud_id}() {{")
        lines.append(f"  var form = FormApp.create({json.dumps(titulo, ensure_ascii=False)});")
        lines.append(f"  form.setDescription({json.dumps(desc, ensure_ascii=False)});")
        lines.append("  form.setIsQuiz(false);")
        lines.append("  form.setCollectEmail(false);")
        lines.append("  form.setAllowResponseEdits(false);")
        lines.append(
            "  form.setConfirmationMessage('Gracias. Sus respuestas alimentan el diagnóstico MDeIA UCCuyo.');"
        )

        for p in aud.get("preguntas", []):
            col = p["columna"]
            texto = p.get("texto", col)
            escala = p.get("escala", "texto")
            req = "true" if p.get("requerida") else "false"
            title = json.dumps(col, ensure_ascii=False)
            help_text = json.dumps(texto, ensure_ascii=False)

            if escala == "texto":
                lines.append(
                    f"  var item = form.addTextItem().setTitle({title}).setHelpText({help_text});"
                )
                lines.append(f"  item.setRequired({req});")
            else:
                if escala == "meta_sede":
                    opts = p.get("opciones") or ["San Juan", "San Luis", "Mendoza"]
                elif escala == "meta_unidad":
                    opts = None  # usa UNIDADES_MDEIA en JS
                else:
                    opts = cfg["escalas"].get(escala, {}).get("opciones", [])
                lines.append(
                    f"  var item = form.addListItem().setTitle({title}).setHelpText({help_text});"
                )
                if escala == "meta_unidad":
                    lines.append(
                        "  item.setChoices(UNIDADES_MDEIA.map(function(o) { return item.createChoice(o); }));"
                    )
                else:
                    opts_js = json.dumps(opts, ensure_ascii=False)
                    lines.append(
                        f"  item.setChoices({opts_js}.map(function(o) {{ return item.createChoice(o); }}));"
                    )
                lines.append(f"  item.setRequired({req});")

        lines.append("  var url = form.getPublishedUrl();")
        lines.append(f"  Logger.log('Formulario {aud['nombre']}: ' + url);")
        lines.append("  return url;")
        lines.append("}")
        lines.append("")

    return "\n".join(lines)


def plantilla_filename(audiencia_id: str) -> str:
    return f"plantilla-encuesta-mdeia-{audiencia_id}.xlsx"
